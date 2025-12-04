import openpyxl 
import os
from datetime import datetime

FILE_NAME = 'data_beasiswa.xlsx'

# MEMBUAT SHEET JIKA BELUM ADA
def create_sheet_if_not_exists(workbook, sheet_name, header=None):
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        if header:
            sheet.append(header)
    return workbook[sheet_name]

# LOAD / CREATE FILE EXCEL
def load_or_create_workbook():
    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook()
        default = workbook.active
        workbook.remove(default)
        workbook.save(FILE_NAME)      
        return workbook
    else:
        return openpyxl.load_workbook(FILE_NAME)

# VALIDASI & MEMECAH KODE BEASISWA
def memisahkan_string(kode):
    if len(kode) != 6:
        print("Kode harus 6 karakter, contoh: B01001")
        return None, None
    
    jenis = kode[0:3]
    nomor = kode[3:6]

    if jenis not in ["B01", "B02", "B03"]:
        print("Jenis kode tidak valid. Gunakan B01, B02, atau B03.")
        return None, None

    return jenis, nomor

# MENENTUKAN JENIS BEASISWA
def tentukan_jenis_beasiswa(jenis_beasiswa):
    jenis_map = {
        "B01": "Beasiswa Pemerintah",
        "B02": "Beasiswa Swasta",
        "B03": "Beasiswa Perguruan Tinggi"
    }
    return jenis_map.get(jenis_beasiswa, "Tidak diketahui")

# Fungsi TAMBAH BEASISWA
def tambah_beasiswa():
    print("\n=== DAFTAR JENIS BEASISWA ===")
    print("B01 = Beasiswa Pemerintah")
    print("B02 = Beasiswa Swasta")
    print("B03 = Beasiswa Perguruan Tinggi")
    print("Contoh Kode lengkap: B01001\n")

    kode = input("Masukkan Kode Beasiswa (B01xxx): ")

    jenis_code, nomor_code = memisahkan_string(kode)
    if jenis_code is None:
        return
    
    nama = input("Masukkan Nama Beasiswa: ")
    pemberi = input("Masukkan Pemberi Beasiswa: ")
    kuota = input("Masukkan Kuota Beasiswa: ")

    if not kuota.isdigit():
        print("Kuota harus angka!")
        return

    workbook = load_or_create_workbook()
    sheet = create_sheet_if_not_exists(
        workbook, 
        'Beasiswa',
        ['Kode', 'Nama', 'Pemberi', 'Jenis', 'Nomor', 'Kuota', 'Status']
    )

    # Cek duplikasi
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == kode:
            print("Kode sudah terdaftar!")
            return

    jenis_nama = tentukan_jenis_beasiswa(jenis_code)

    sheet.append([kode, nama, pemberi, jenis_nama, nomor_code, kuota, 'Tersedia'])
    workbook.save(FILE_NAME)
    print("Beasiswa berhasil ditambahkan.")

# Fungsi TAMPIL BEASISWA
def tampil_beasiswa():
    if not os.path.exists(FILE_NAME):
        print("Data tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Beasiswa' not in workbook.sheetnames:
        print("Sheet Beasiswa belum ada.")
        return

    sheet = workbook['Beasiswa']

    if sheet.max_row == 1:
        print("Belum ada data.")
        return

    print("\n=== DAFTAR BEASISWA ===")
    header = ["Kode", "Nama", "Pemberi", "Jenis", "Nomor", "Kuota", "Status"]
    print("{:<10} {:<25} {:<20} {:<30} {:<8} {:<8} {:<10}".format(*header))

    for row in sheet.iter_rows(min_row=2, values_only=True):
        print("{:<10} {:<25} {:<20} {:<30} {:<8} {:<8} {:<10}".format(*row))

# Fungsi EDIT BEASISWA
def edit_beasiswa():
    kode = input("Masukkan kode beasiswa yang ingin diedit: ")

    if not os.path.exists(FILE_NAME):
        print("File tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Beasiswa' not in workbook.sheetnames:
        print("Sheet tidak ada.")
        return

    sheet = workbook['Beasiswa']

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == kode:
            print("Data ditemukan. Kosongkan jika tidak ingin mengubah.")

            nama = input("Nama baru: ")
            pemberi = input("Pemberi baru: ")
            kuota = input("Kuota baru: ")

            if nama:
                row[1].value = nama
            
            if pemberi:
                row[2].value = pemberi

            if kuota.isdigit():
                row[5].value = kuota

            workbook.save(FILE_NAME)
            print("Beasiswa berhasil diedit.")
            return

    print("Beasiswa tidak ditemukan.")

# Fungsi HAPUS BEASISWA
def hapus_beasiswa():
    kode = input("Masukkan kode beasiswa yang ingin dihapus: ")

    if not os.path.exists(FILE_NAME):
        print("Data tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)

    if 'Beasiswa' not in workbook.sheetnames:
        print("Sheet tidak ada.")
        return

    sheet = workbook['Beasiswa']

    for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row[0].value == kode:
            konfirm = input("Yakin hapus? (y/n): ").lower()
            if konfirm == 'y':
                sheet.delete_rows(idx)
                workbook.save(FILE_NAME)
                print("Beasiswa berhasil dihapus.")
            else:
                print("Dibatalkan.")
            return

    print("Kode tidak ditemukan.")

# MENU BEASISWA
def menu_beasiswa():
    while True:
        print("\n=== MENU BEASISWA ===")
        print("1. Tambah Beasiswa")
        print("2. Tampil Beasiswa")
        print("3. Edit Beasiswa")
        print("4. Hapus Beasiswa")
        print("5. Kembali ke Menu Utama")

        pilihan = input("Pilih menu: ")

        if pilihan == '1':
            tambah_beasiswa()
        elif pilihan == '2':
            tampil_beasiswa()
        elif pilihan == '3':
            edit_beasiswa()
        elif pilihan == '4':
            hapus_beasiswa()
        elif pilihan == '5':
            break
        else:
            print("Pilihan tidak valid.")
