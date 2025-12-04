import openpyxl 
import os

FILE_NAME = 'data_beasiswa.xlsx'

def create_sheet_if_not_exists(workbook, sheet_name, header=None):
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        if header:
            sheet.append(header)
    return workbook[sheet_name]

def load_or_create_workbook():
    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook()
        default = workbook.active
        workbook.remove(default)
        workbook.save(FILE_NAME)
        return workbook
    else:
        return openpyxl.load_workbook(FILE_NAME)

# Fungsi TAMBAH SISWA 
def tambah_siswa():
    nisn_siswa = input("Masukkan NISN Siswa: ")
    
    if not nisn_siswa.isdigit():
        print("NISN harus angka!")
        return

    nama_siswa = input("Masukkan Nama Siswa: ")
    no_hp = input("Masukkan No. HP Siswa:  ")
    alamat = input("Masukkan Alamat Siswa: ")

    workbook = load_or_create_workbook()
    sheet = create_sheet_if_not_exists(
        workbook, 
        'Siswa', 
        ['NISN','Nama Siswa','No HP','Alamat']
    )

    # Cek duplikasi
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == nisn_siswa:
            print("GAGAL! NISN sudah terdaftar!")
            return

    sheet.append([nisn_siswa, nama_siswa, no_hp, alamat])
    workbook.save(FILE_NAME)
    print("Data Siswa berhasil ditambahkan.")

# Fungsi TAMPIL SISWA 
def tampil_siswa():
    if not os.path.exists(FILE_NAME):
        print("File data siswa tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet Siswa belum ada.")
        return

    sheet = workbook['Siswa']

    if sheet.max_row == 1:
        print("Belum ada data siswa.")
        return

    print("\n=== DAFTAR SISWA BERHAK BEASISWA ===")
    print("{:<15} {:<20} {:<15} {:<30}".format("NISN","Nama","No HP","Alamat"))

    for row in sheet.iter_rows(min_row=2, values_only=True):
        print("{:<15} {:<20} {:<15} {:<30}".format(*row))

# Fungsi EDIT SISWA
def edit_siswa():
    nisn_siswa = input("Masukkan NISN Siswa yang ingin diedit: ")

    if not os.path.exists(FILE_NAME):
        print("File data siswa tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet Siswa belum ada.")
        return

    sheet = workbook['Siswa']

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == nisn_siswa:
            print("\nData ditemukan. Kosongkan jika tidak ingin mengubah.")

            nama_baru = input("Nama baru: ")
            nohp_baru = input("No HP baru : ")
            alamat_baru = input("Alamat baru: ")

            if nama_baru: row[1].value = nama_baru
            if nohp_baru: row[2].value = nohp_baru
            if alamat_baru: row[3].value = alamat_baru

            workbook.save(FILE_NAME)
            print("Data Siswa berhasil diedit.")
            return

    print("Data tidak ditemukan.")

# Fungsi HAPUS SISWA
def hapus_siswa():
    nisn_siswa = input("Masukkan NISN Siswa yang ingin dihapus: ")

    if not os.path.exists(FILE_NAME):
        print("File tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet Siswa belum ada.")
        return

    sheet = workbook['Siswa']

    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row[0].value == nisn_siswa:
            konfirmasi = input("Yakin ingin hapus data? (y/n): ").lower()
            if konfirmasi != "y":
                print("Penghapusan dibatalkan.")
                return

            sheet.delete_rows(row_index)
            workbook.save(FILE_NAME)
            print("Data berhasil dihapus.")
            return

    print("Data tidak ditemukan.")

# MENU SISWA
def menu_siswa():
    while True:
        print("\n=== DATA SISWA BERHAK BEASISWA ===")
        print("1. Tambah Data Siswa")
        print("2. Tampil Data Siswa")
        print("3. Edit Data Siswa")
        print("4. Hapus Data Siswa")
        print("5. Kembali ke Menu Utama")

        pilihan = input("Pilih menu: ")

        if pilihan == '1':
            tambah_siswa()
        elif pilihan == '2':
            tampil_siswa()
        elif pilihan == '3':
            edit_siswa()
        elif pilihan == '4':
            hapus_siswa()
        elif pilihan == '5':
            break
        else:
            print("Pilihan tidak valid.")