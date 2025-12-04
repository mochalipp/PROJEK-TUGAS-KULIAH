import openpyxl
import os
from datetime import datetime

FILE_NAME = 'data_beasiswa.xlsx'

def create_sheet_if_not_exists(workbook, sheet_name, header=None):
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        if header:
            sheet.append(header)
    return workbook[sheet_name]

#FUNGSI PEMBERIAN BEASISWA
def pemberian_beasiswa():
    nisn = input("Masukkan NISN Siswa: ")
    kode = input("Masukkan Kode Beasiswa: ")
    tanggal = datetime.today().strftime("%Y-%m-%d")
    
    # Validasi input jika user meninputkan kode tidak alfanumerik dengan huruf kapital.
    if not kode.isalnum() or not any(kode.isupper() for kode in kode):
        print("Kode beasiswa harus alfanumerik dengan huruf kapital.")
    else : 
        print("Kode beasiswa tidak valid.")
    # Cek keberadaan file pada Excel pada Workbook 
    if not os.path.exists(FILE_NAME):
        print("File data tidak ditemukan.")
        return
    
    # Load workbook untuk validasi data siswa dan beasiswa
    workbook = openpyxl.load_workbook(FILE_NAME)

    # Cek keberadaan sheet Siswa dan Beasiswa jika tidak ada maka tampilkan pesan
    if 'Siswa' not in workbook.sheetnames or 'Beasiswa' not in workbook.sheetnames:
        print("Data siswa atau beasiswa tidak ada.")
        return

    siswa_sheet = workbook['Siswa']
    bea_sheet = workbook['Beasiswa']

    # Validasi siswa untuk memastikan NISN terdaftar pada didalam sheet Siswa
    siswa_valid = any(row[0].value == nisn for row in siswa_sheet.iter_rows(min_row=2))
    if not siswa_valid:
        print("Siswa tidak terdaftar.")
        return

    # Validasi beasiswa dan kuota beasiswa tersedia atau tidak 
    bea_row = None
    for row in bea_sheet.iter_rows(min_row=2):
        if row[0].value == kode:
            bea_row = row
            break
    
    # Jika beasiswa tidak ditemukan pada row maka tampilkan pesan
    if bea_row is None:
        print("Beasiswa tidak ditemukan.")
        return

    kuota = int(bea_row[5].value)

    if kuota <= 0:
        print("Kuota beasiswa habis.")
        return

    # Sheet pemberian beasiswa lalu dibuat jika belum ada dalam workbook
    pemberian_sheet = create_sheet_if_not_exists(
        workbook, 
        'Pemberian', 
        ['NISN', 'Kode Beasiswa', 'Tanggal']
    )

    # Cek duplikasi pemberian beasiswa dan tampilkan pesan jika sudah pernah diberikan sebelumnya 
    for row in pemberian_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == nisn and row[1] == kode:
            print("Siswa sudah pernah menerima beasiswa ini.")
            return

    # Simpan data pada sheet pemberian beasiswa
    pemberian_sheet.append([nisn, kode, tanggal])

    # Kurangi kuota
    kuota -= 1
    bea_row[5].value = kuota
    bea_row[6].value = "Habis" if kuota == 0 else "Tersedia"

    workbook.save(FILE_NAME)
    print("Beasiswa berhasil diberikan.")

#FUNGSI PENCABUTAN BEASISWA 
def pencabutan_beasiswa():
    nisn = input("Masukkan NISN Siswa: ")
    kode = input("Masukkan Kode Beasiswa: ")

    if not os.path.exists(FILE_NAME):
        print("File data tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)

    if 'Pemberian' not in workbook.sheetnames:
        print("Belum ada data pemberian.")
        return

    if 'Beasiswa' not in workbook.sheetnames:
        print("Data beasiswa tidak ditemukan.")
        return

    pemberian_sheet = workbook['Pemberian']
    bea_sheet = workbook['Beasiswa']

    # Cari data pemberian
    target_row = None
    for row in pemberian_sheet.iter_rows(min_row=2):
        if row[0].value == nisn and row[1].value == kode:
            target_row = row
            break

    if target_row is None:
        print("Data pemberian tidak ditemukan.")
        return

    # Hapus baris pemberian
    row_number = target_row[0].row
    pemberian_sheet.delete_rows(row_number)

    # Kembalikan kuota
    for row in bea_sheet.iter_rows(min_row=2):
        if row[0].value == kode:
            kuota = int(row[5].value)
            row[5].value = kuota + 1
            row[6].value = "Tersedia"
            break

    # Tambah ke histori
    histori_sheet = create_sheet_if_not_exists(
        workbook,
        'Pencabutan',
        ['NISN', 'Kode Beasiswa', 'Tanggal Pencabutan']
    )

    histori_sheet.append([nisn, kode, datetime.today().strftime("%Y-%m-%d")])

    workbook.save(FILE_NAME)
    print("Pencabutan beasiswa berhasil.")

#FUNGSI TAMPIL DATA PEMBERIAN
def tampil_data_pemberian():
    if not os.path.exists(FILE_NAME):
        print("File tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)

    if 'Pemberian' not in workbook.sheetnames:
        print("Belum ada data pemberian.")
        return

    sheet = workbook['Pemberian']

    print("\n=== Data Pemberian Beasiswa ===")
    print("{:<15} {:<20} {:<15}".format("NISN", "Kode Beasiswa", "Tanggal"))

    for row in sheet.iter_rows(min_row=2, values_only=True):
        print("{:<15} {:<20} {:<15}".format(*row))

#FUNSGSI TAMPIL HISTORI PENCABUTAN
def tampil_history_pencabutan():
    if not os.path.exists(FILE_NAME):
        print("File tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)

    if 'Histori_Pencabutan' not in workbook.sheetnames:
        print("Belum ada histori pencabutan.")
        return

    sheet = workbook['Histori_Pencabutan']

    print("\n=== Histori Pencabutan Beasiswa ===")
    print("{:<15} {:<20} {:<15}".format("NISN", "Kode Beasiswa", "Tanggal"))

    for row in sheet.iter_rows(min_row=2, values_only=True):
        print("{:<15} {:<20} {:<15}".format(*row))

#FUNGSI MENU PEMBERIAN
def menu_pemberian():
    while True:
        print("\n=== Menu Pemberian Beasiswa ===")
        print("1. Tambahkan Pemberian Beasiswa")
        print("2. Tampilkan Pemberian Beasiswa")
        print("3. Cabut Pemberian Beasiswa")
        print("4. Tampilkan Histori Pencabutan")
        print("5. Kembali ke Menu Utama")

        pilihan = input("Pilih menu: ")

        if pilihan == '1':
            pemberian_beasiswa()
        elif pilihan == '2':
            tampil_data_pemberian()
        elif pilihan == '3':
            pencabutan_beasiswa()
        elif pilihan == '4':
            tampil_history_pencabutan()
        elif pilihan == '5':
            break
        else:
            print("Pilihan tidak valid.")