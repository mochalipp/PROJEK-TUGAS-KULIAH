import openpyxl 
import os

FILE_NAME = 'data_beasiswa.xlsx'

# Fungsi mendapatkan nama siswa berdasarkan NISN 
def get_nama_siswa(workbook, nisn):
    if 'Siswa' not in workbook.sheetnames:
        return "-"
    sheet = workbook['Siswa']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == nisn:
            return row[1]
    return "Nama Siswa"

# fungsi mendapatkan nama beasiswa berdasarkan kode
def get_nama_beasiswa(workbook, kode):
    if 'Beasiswa' not in workbook.sheetnames:
        return "-"
    sheet = workbook['Beasiswa']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == kode:
            return row[1]
    return "Nama Beasiswa"

# LAPORAN DISTRIBUSI BEASISWA
def tampil_laporan_distribusi():
    if not os.path.exists(FILE_NAME):
        print("File data beasiswa tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)

    if 'Pemberian' not in workbook.sheetnames:
        print("Belum ada data pemberian beasiswa.")
        return

    sheet = workbook['Pemberian']

    if sheet.max_row == 1:
        print("Belum ada data distribusi beasiswa.")
        return

    print("\n=== LAPORAN DISTRIBUSI BEASISWA ===")
    print("-" * 70)

    for nisn, kode, tanggal in sheet.iter_rows(min_row=1, values_only=True):
        nama_siswa = get_nama_siswa(workbook, nisn)
        nama_bea = get_nama_beasiswa(workbook, kode)

        print("{:<15} {:<20} {:<20} {:<15}".format(nisn, nama_siswa, nama_bea, tanggal))
        print("-" * 70)

# MENU LAPORAN
def menu_laporan():
    while True:
        print("\n === MENU LAPORAN ===")
        print("1. Laporan Distribusi Beasiswa")
        print("2. Kembali ke Menu Utama")

        pilihan = input("Pilih menu: ")

        if pilihan == '1':
            tampil_laporan_distribusi()
        elif pilihan == '2':
            break
        else:
            print("Pilihan tidak valid.")